import psycopg2
from openpyxl import load_workbook


def comisionar(Ano, Mes, FIniA, FIniB, FFinA, FFinB, progressBar, taskInfo):
    # Connect to an existing database
    try:
        conn = psycopg2.connect(dbname="Lab_Argos", user="postgres", host="192.168.1.150", port="5433",
                                password="250996")
    except:
        print("I am unable to connect to the database")

    # Open a cursor to perform database operations
    cur = conn.cursor()

    progressBar.setValue(1)
    taskInfo.setText("Conectado a la Base de Datos")


    progressBar.setValue(2)
    taskInfo.setText("Inicializando Informe...")

    wb = load_workbook(filename="//labnas/Gerencia/Comisiones/Comisiones.xlsx")

    ws = wb["PORTADA"]
    ws.cell(row=1, column=2, value=Mes)
    ws.cell(row=2, column=2, value=Ano)
    ws.cell(row=5, column=2, value=FIniB)
    ws.cell(row=5, column=3, value=FFinB)
    ws.cell(row=6, column=2, value=FIniA)
    ws.cell(row=6, column=3, value=FFinA)

    # VENTAS A
    cur.execute("""select numero_de_venta,fecha,total,viajante,sum(valorteorico) as teorico,sum(valorreal) as real from
(SELECT 
  (articulos_de_ventas.cantidad * articulos_de_ventas.precio_unitario) - (articulos_de_ventas.cantidad * articulos_de_ventas.precio_unitario * articulos_de_ventas.descuento/100) as valorreal,
  articulos_de_ventas.cantidad * productos.precio_venta as valorteorico,
  ventas.numero_de_venta, 
  ventas.total,
  ventas.fecha, 
  ventas.viajante
FROM 
	productos
inner join 
	articulos_de_ventas on articulos_de_ventas.codigo=productos.codigo
inner join
	ventas on articulos_de_ventas.numero_de_venta=ventas.numero_de_venta
WHERE 
  ventas.fecha BETWEEN '""" + FIniA + """"' AND '""" + FFinA + """"' and ventas.tipo_de_venta='V') as test
group by numero_de_venta,fecha,total,viajante
order by
  fecha ASC, 
  numero_de_venta ASC;""")

    rows = cur.fetchall()

    ws = wb["VentasA"]

    i = 0
    for row in rows:
        ws.cell(row=10 + i, column=2, value=row[0])
        ws.cell(row=10 + i, column=3, value=row[1])
        ws.cell(row=10 + i, column=4, value=row[2])
        ws.cell(row=10 + i, column=5, value=row[3])
        ws.cell(row=10 + i, column=6, value=row[4])
        ws.cell(row=10 + i, column=7, value=row[5])
        i = i + 1

    progressBar.setValue(3)
    taskInfo.setText("Procesadas Vantas A...")


    # VENTAS B
    cur.execute("""select factura_numero,fecha,total,viajante,sum(valorteorico) as teorico,sum(valorreal) as real from
(SELECT 
  (articulos_de_facturas_de_venta.cantidad * articulos_de_facturas_de_venta.precio_unitario) - (articulos_de_facturas_de_venta.cantidad * articulos_de_facturas_de_venta.precio_unitario * articulos_de_facturas_de_venta.descuento/100) as valorreal,
  articulos_de_facturas_de_venta.cantidad * productos.precio_venta as valorteorico,
  facturas_de_venta.factura_numero, 
  facturas_de_venta.total,
  facturas_de_venta.fecha, 
  facturas_de_venta.viajante
FROM 
	productos
inner join 
	articulos_de_facturas_de_venta on articulos_de_facturas_de_venta.codigo=productos.codigo
inner join
	facturas_de_venta on articulos_de_facturas_de_venta.factura_numero=facturas_de_venta.factura_numero
WHERE 
  facturas_de_venta.fecha BETWEEN '""" + FIniB + """"' AND '""" + FFinB + """"' and facturas_de_venta.tipo_de_factura like 'F%' and articulos_de_facturas_de_venta.punto_venta=3) as test
group by factura_numero,fecha,total,viajante
order by
  fecha ASC, 
  factura_numero ASC;""")

    rows = cur.fetchall()

    ws = wb["VentasB"]

    i = 0
    for row in rows:
        ws.cell(row=10 + i, column=2, value=row[0])
        ws.cell(row=10 + i, column=3, value=row[1])
        ws.cell(row=10 + i, column=4, value=row[2])
        ws.cell(row=10 + i, column=5, value=row[3])
        ws.cell(row=10 + i, column=6, value=row[4])
        ws.cell(row=10 + i, column=7, value=row[5])
        i = i + 1

    progressBar.setValue(4)
    taskInfo.setText("Procesadas Vantas B...")

    # COBRANZAS A
    cur.execute("""SELECT 
  imputaciones_de_pagos_de_ventas.numero_de_pago, 
  pagos_de_ventas.fecha, 
  imputaciones_de_pagos_de_ventas.numero_de_venta, 
  pagos_de_ventas.fecha-ventas.fecha, 
  imputaciones_de_pagos_de_ventas.importe_imputado,
  ventas.viajante,
  clientes.provincia
FROM public.imputaciones_de_pagos_de_ventas
inner join public.ventas on public.ventas.numero_de_venta=public.imputaciones_de_pagos_de_ventas.numero_de_venta
inner join public.pagos_de_ventas on public.pagos_de_ventas.numero_de_pago=public.imputaciones_de_pagos_de_ventas.numero_de_pago
inner join public.clientes on public.clientes.cod_cliente=public.pagos_de_ventas.cod_cliente

WHERE 
  imputaciones_de_pagos_de_ventas.numero_de_pago = pagos_de_ventas.numero_de_pago AND
  ventas.numero_de_venta = imputaciones_de_pagos_de_ventas.numero_de_venta AND
  ventas.tipo_de_venta = 'V' AND
  pagos_de_ventas.fecha BETWEEN '""" + FIniA + """"' AND '""" + FFinA + """"';""")

    rows = cur.fetchall()

    ws = wb["CobranzasA"]

    i = 0
    for row in rows:
        ws.cell(row=12 + i, column=2, value=row[0])
        ws.cell(row=12 + i, column=3, value=row[1])
        ws.cell(row=12 + i, column=4, value=row[2])
        ws.cell(row=12 + i, column=5, value=row[3])
        ws.cell(row=12 + i, column=6, value=row[4])
        ws.cell(row=12 + i, column=7, value=row[5])
        ws.cell(row=12 + i, column=8, value=row[6])
        i = i + 1

    progressBar.setValue(5)
    taskInfo.setText("Procesadas Cobranzas A...")

    # COBRANZAS B
    cur.execute("""SELECT 
  imputaciones_de_recibos_de_clientes.numero_de_recibo, 
  recibos_de_clientes.fecha, 
  imputaciones_de_recibos_de_clientes.factura_numero, 
  recibos_de_clientes.fecha-facturas_de_venta.fecha, 
  imputaciones_de_recibos_de_clientes.importe_imputado,
  facturas_de_venta.viajante,
  recibos_de_clientes.provincia
FROM 
  public.imputaciones_de_recibos_de_clientes, 
  public.facturas_de_venta, 
  public.recibos_de_clientes
WHERE 
  imputaciones_de_recibos_de_clientes.numero_de_recibo = recibos_de_clientes.numero_de_recibo AND
  facturas_de_venta.factura_numero = imputaciones_de_recibos_de_clientes.factura_numero AND
  facturas_de_venta.tipo_de_factura = imputaciones_de_recibos_de_clientes.tipo_de_factura AND
  facturas_de_venta.tipo_de_factura like 'F%' AND
  facturas_de_venta.punto_venta = imputaciones_de_recibos_de_clientes.punto_venta AND
  recibos_de_clientes.fecha BETWEEN '""" + FIniB + """"' AND '""" + FFinB + """"';
""")

    rows = cur.fetchall()

    ws = wb["CobranzasB"]

    i = 0
    for row in rows:
        ws.cell(row=12 + i, column=2, value=row[0])
        ws.cell(row=12 + i, column=3, value=row[1])
        ws.cell(row=12 + i, column=4, value=row[2])
        ws.cell(row=12 + i, column=5, value=row[3])
        ws.cell(row=12 + i, column=6, value=row[4])
        ws.cell(row=12 + i, column=7, value=row[5])
        ws.cell(row=12 + i, column=8, value=row[6])
        i = i + 1

    progressBar.setValue(6)
    taskInfo.setText("Procesadas Cobranzas B...")

    # CREDITO A
    cur.execute("""SELECT 
  numero_de_venta, 
  fecha, 
  total,
  viajante  
FROM 
  ventas
WHERE 
  tipo_de_venta LIKE 'NC%' AND 
  fecha BETWEEN '""" + FIniA + """"' AND '""" + FFinA + """"'
ORDER BY
  fecha ASC, 
  numero_de_venta ASC;""")

    rows = cur.fetchall()

    ws = wb["DevolucionesA"]

    i = 0
    for row in rows:
        ws.cell(row=10 + i, column=2, value=row[0])
        ws.cell(row=10 + i, column=3, value=row[1])
        ws.cell(row=10 + i, column=4, value=row[2])
        ws.cell(row=10 + i, column=5, value=row[3])
        i = i + 1

    progressBar.setValue(7)
    taskInfo.setText("Procesados Creditos A...")

    # CREDITO B
    cur.execute("""SELECT 
  facturas_de_venta.factura_numero, 
  facturas_de_venta.fecha, 
  facturas_de_venta.total,
  facturas_de_venta.viajante
FROM 
  public.facturas_de_venta
WHERE 
  facturas_de_venta.tipo_de_factura LIKE 'NC%' AND 
  fecha BETWEEN '""" + FIniA + """"' AND '""" + FFinA + """"'
ORDER BY
  facturas_de_venta.fecha ASC, 
  facturas_de_venta.factura_numero ASC;
""")

    rows = cur.fetchall()

    ws = wb["DevolucionesB"]

    i = 0
    for row in rows:
        ws.cell(row=10 + i, column=2, value=row[0])
        ws.cell(row=10 + i, column=3, value=row[1])
        ws.cell(row=10 + i, column=4, value=row[2])
        ws.cell(row=10 + i, column=5, value=row[3])
        i = i + 1

    progressBar.setValue(8)
    taskInfo.setText("Procesados Creditos B...")

    cur.close()
    conn.close()

    progressBar.setValue(9)
    taskInfo.setText("Cerrando BAse de Datos...")

    wb.save(filename="//labnas/Gerencia/Comisiones/Comisiones"+str(Ano)+str(Mes)+".xlsx")
    wb.close()

    progressBar.setValue(10)
    taskInfo.setText("FINALIZADO...")

    print("FIN")